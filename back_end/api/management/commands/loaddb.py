from django.core.management.base import BaseCommand, CommandError
from api.models import *
from datetime import date

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print(f'กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A1'].value)

        print(f'count = {count}')
        data = []
        for i in range(count):  # 0,1,2,3
            # print(f'i = {i}')
            sheet_values = [
                ws[f'{chr(65+j)}{3+i}'].value for j in range(len(column_names))
            ]
            data.append(
                dict((k, v) for k, v in zip(column_names, sheet_values)))

        return data

    def handle(self, *args, **options):

        from openpyxl import load_workbook
        filename = "xlsx/loaddata.xlsx"
        wb = load_workbook(filename, data_only=True)

        for b in self.load(wb, 'BloodType', ['id', 'bloodtype']):
            print("data  =  ",b)
            q = BloodType(**b)
            q.save()
        print("seve...Blood")
        for n in self.load(wb, 'NakSus', ['id', 'naksus']):
            print("data =  ", n)
            q = NakSus(**n)
            q.save()
        # print("seve...Blood")

        for d in self.load(wb, 'RaSi', ['id', 'rasi']):
            print("RaSi =  ", d)
            # (**w).save()
            q = RaSi(**d)
            q.save()
        for d in self.load(wb, 'Gender', ['id', 'gender']):
            print("Gender =  ", d)
            # (**w).save()
            q = Gender(**d)
            q.save()
        for d in self.load(wb, 'Testes', ['id', 'testes']):
            print("Testes =  ", d)
            # (**w).save()
            q = Testes(**d)
            q.save()
        for d in self.load(wb, 'DaysOfWeek', ['id', 'daysofweek']):
            print("DaysofWeek =  ", d)
            # (**w).save()
            q = DaysOfWeek(**d)
            q.save()

        for d in self.load(wb, 'Member', [
                'id', 'username', 'email', 'password', 'first_name',
                'last_name', 'birthday', 'age', 'dayofbirth', 'rasi',
                'bloodtype', 'naksus', 'gender', 'testes', 'profileurl',
                'discription', 'characterneed', 'values', 'created_at',
                'updated_at'
        ]):

            dayofbirth = DaysOfWeek.objects.get(pk=d['dayofbirth'])
            d.pop('dayofbirth', None)

            rasi = RaSi.objects.get(pk=d['rasi'])
            d.pop('rasi', None)

            bloodType = BloodType.objects.get(pk=d['bloodtype'])
            d.pop('bloodtype', None)

            nakSus = NakSus.objects.get(pk=d['naksus'])
            d.pop('naksus', None)

            genDer = Gender.objects.get(pk=d['gender'])
            d.pop('gender', None)

            testes = Testes.objects.get(pk=d['testes'])
            d.pop('testes', None)


            q = Member(**d)


            q.dayofbirth = dayofbirth
            q.rasi = rasi
            q.bloodtype = bloodType
            q.naksus = nakSus
            q.gender = genDer
            q.testes = testes
            # q.updated_at = updateAt

            q.save()

        for d in self.load(wb, 'Conversation', [
                'id', 'member', 'message', 'block', 'rejected',
                'reviewe_value', 'joined_at', 'updated_at'
        ]):
            print("Conversation =  ", d)
            # (**w).save()
            member = Member.objects.get(pk=d['member'])
            d.pop('member', None)


            q = Conversation(**d)
            q.member = member
            q.save()

        for d in self.load(wb, 'Goldmember', [
               'id' ,'goldmember' ,'conversation'

        ]):
            print("Goldmember =  ", d)
            # (**w).save()
            goldmember = Member.objects.get(pk=d['goldmember'])
            d.pop('goldmember', None)
            conversation = Conversation.objects.get(pk=d['conversation'])
            d.pop('conversation', None)

            q = Goldmember(**d)

            q.goldmember = goldmember
            q.conversation = conversation
            q.save()
