from members.signals import save_profile
from django.core.management.base import BaseCommand, CommandError
from members.models import *
# from django.contrib.auth.hashers import make_password, HASHERS
from django.contrib.auth.models import  User

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print(f'กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A1'].value)

        print(f'count = {count}')
        data = []
        for i in range(count):  # 0,1,2,3
            # print(f'i = {i}')
            sheet_values = [
                ws[f'{chr(65+j)}{3+i}'].value for j in range(len(column_names))
            ]
            data.append(
                dict((k, v) for k, v in zip(column_names, sheet_values)))

        return data

    def handle(self, *args, **options):

        
        filename = "xlsx/loaddata.xlsx"
        wb = load_workbook(filename, data_only=True)
        # for user in Member.objects.all():
        #     user.set_password(user.password)
        #     user.save()

        for b in self.load(wb, 'Member', ['id', 'first_name', 'last_name', 'username', 'email',]):
            print("User  =  ",b)
            q = Member(**b)
            q.save()
        print("seve...User")
        print("____________________________")

        