from django.core.management.base import BaseCommand, CommandError
from api.models import RaSi
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def handle(self, *args, **optiSons):
        # pass
       
        file = "xlsx/loaddata.xlsx"
        wb = load_workbook(file, read_only=True)
        # datas = wb.get_sheet_names()[2]
        # ws = wb.get_sheet_by_name(datas)
        
    
        ws = wb['Data1']
        count = int(ws['A1'].value)
        print(f'User count = {count}')
        print(count)
        
        for i in range(count):
            print(i)
            
        row = [ ws[f'{c}2'].value for c in 'AB' ]
        print( row )
        # for i in range(count): # 0,1,2,3
        #     q = User(**{
        #         "user_name" : str(ws[f'B{3+i}'].value),
        #         "user_lastname" : str(ws[f'C{3+i}'].value),
        #         "phone" : str(ws[f'D{3+i}'].value),
        #         "email" : str(ws[f'E{3+i}'].value),
        #         "username" : str(ws[f'F{3+i}'].value),
        #         "password" : str(ws[f'G{3+i}'].value),
        #         "avatar": str(ws[f'H{3+i}'].value), #รูปผู้ใช้งาน
        #         "lat"  : int(ws[f'I{3+i}'].value),
        #         "lng" : int(ws[f'J{3+i}'].value),

        #         })
        #     q.save()
